"""
Export utilities for Turbo Air Equipment Viewer
Handles Excel and PDF generation for quotes
"""

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from typing import Dict, List
import io
from PIL import Image

# Color constants
TURBO_BLUE = '#20429c'
TURBO_RED = '#d3242b'

def generate_excel_quote(quote_data: Dict, items_df: pd.DataFrame, client_data: Dict) -> io.BytesIO:
    """Generate Excel quote with product images"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Quote {quote_data['quote_number']}"
    
    # Define styles
    header_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    blue_fill = PatternFill(start_color="20429C", end_color="20429C", fill_type="solid")
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header Section
    ws.merge_cells('A1:H2')
    ws['A1'] = 'TURBO AIR EQUIPMENT QUOTE'
    ws['A1'].font = header_font
    ws['A1'].fill = blue_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Quote Information
    ws['A4'] = 'Quote Number:'
    ws['B4'] = quote_data['quote_number']
    ws['B4'].font = bold_font
    
    ws['A5'] = 'Date:'
    ws['B5'] = datetime.now().strftime('%B %d, %Y')
    
    ws['A6'] = 'Valid Until:'
    ws['B6'] = (datetime.now() + pd.Timedelta(days=30)).strftime('%B %d, %Y')
    
    # Client Information
    ws['D4'] = 'Client:'
    ws['E4'] = client_data.get('company', '')
    ws['E4'].font = bold_font
    
    ws['D5'] = 'Contact:'
    ws['E5'] = client_data.get('contact_name', '')
    
    ws['D6'] = 'Email:'
    ws['E6'] = client_data.get('contact_email', '')
    
    ws['D7'] = 'Phone:'
    ws['E7'] = client_data.get('contact_number', '')
    
    # Product Table Headers
    start_row = 10
    headers = ['Image', 'SKU', 'Description', 'Specifications', 'Quantity', 'Unit Price', 'Total']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Image
    ws.column_dimensions['B'].width = 15  # SKU
    ws.column_dimensions['C'].width = 30  # Description
    ws.column_dimensions['D'].width = 40  # Specifications
    ws.column_dimensions['E'].width = 10  # Quantity
    ws.column_dimensions['F'].width = 12  # Unit Price
    ws.column_dimensions['G'].width = 12  # Total
    
    # Add product rows
    current_row = start_row + 1
    
    for _, item in items_df.iterrows():
        # Set row height for image
        ws.row_dimensions[current_row].height = 60
        
        # Try to add product image
        image_path = f"pdf_screenshots/{item['sku']}/page_1.png"
        if os.path.exists(image_path):
            try:
                img = XLImage(image_path)
                img.width = 80
                img.height = 60
                img.anchor = f'A{current_row}'
                ws.add_image(img)
            except:
                ws.cell(row=current_row, column=1, value="No Image")
        else:
            ws.cell(row=current_row, column=1, value="No Image")
        
        # Product details
        ws.cell(row=current_row, column=2, value=item['sku']).border = thin_border
        ws.cell(row=current_row, column=3, value=item.get('description', '')).border = thin_border
        
        # Specifications (combine key specs)
        specs = []
        if item.get('dimensions'):
            specs.append(f"Dimensions: {item['dimensions']}")
        if item.get('voltage'):
            specs.append(f"Voltage: {item['voltage']}")
        if item.get('temperature_range'):
            specs.append(f"Temp Range: {item['temperature_range']}")
        
        ws.cell(row=current_row, column=4, value='\n'.join(specs)).border = thin_border
        ws.cell(row=current_row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Quantity and prices
        ws.cell(row=current_row, column=5, value=item['quantity']).border = thin_border
        ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='center')
        
        unit_price = ws.cell(row=current_row, column=6, value=item['price'])
        unit_price.border = thin_border
        unit_price.number_format = '$#,##0.00'
        
        total_price = ws.cell(row=current_row, column=7, value=item['price'] * item['quantity'])
        total_price.border = thin_border
        total_price.number_format = '$#,##0.00'
        
        current_row += 1
    
    # Add totals
    current_row += 1
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws.cell(row=current_row, column=1, value='TOTAL:').font = bold_font
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')
    
    total_cell = ws.cell(row=current_row, column=7, value=quote_data['total_amount'])
    total_cell.font = bold_font
    total_cell.number_format = '$#,##0.00'
    total_cell.fill = gray_fill
    total_cell.border = thin_border
    
    # Terms and conditions
    current_row += 3
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws.cell(row=current_row, column=1, value='Terms and Conditions:').font = bold_font
    
    current_row += 1
    terms = [
        '• Prices are valid for 30 days from quote date',
        '• Delivery time: 2-4 weeks from order confirmation',
        '• Payment terms: Net 30 days',
        '• All prices are in USD and exclude taxes',
        '• Installation not included unless specified'
    ]
    
    for term in terms:
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws.cell(row=current_row, column=1, value=term).font = normal_font
        current_row += 1
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def generate_pdf_quote(quote_data: Dict, items_df: pd.DataFrame, client_data: Dict) -> io.BytesIO:
    """Generate PDF quote with professional layout"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor(TURBO_BLUE),
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor(TURBO_BLUE),
        spaceAfter=12
    )
    
    # Title
    elements.append(Paragraph("TURBO AIR EQUIPMENT QUOTE", title_style))
    elements.append(Spacer(1, 0.25*inch))
    
    # Quote and Client Information Table
    info_data = [
        ['Quote Number:', quote_data['quote_number'], 'Client:', client_data.get('company', '')],
        ['Date:', datetime.now().strftime('%B %d, %Y'), 'Contact:', client_data.get('contact_name', '')],
        ['Valid Until:', (datetime.now() + pd.Timedelta(days=30)).strftime('%B %d, %Y'), 
         'Email:', client_data.get('contact_email', '')],
        ['', '', 'Phone:', client_data.get('contact_number', '')]
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (2, 0), (2, -1), 'Helvetica-Bold', 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Products Header
    elements.append(Paragraph("Products", heading_style))
    
    # Product Table
    table_data = [['SKU', 'Description', 'Qty', 'Unit Price', 'Total']]
    
    for _, item in items_df.iterrows():
        table_data.append([
            item['sku'],
            Paragraph(item.get('description', '')[:50] + '...', styles['Normal']),
            str(item['quantity']),
            f"${item['price']:,.2f}",
            f"${item['price'] * item['quantity']:,.2f}"
        ])
    
    # Add total row
    table_data.append(['', '', '', 'TOTAL:', f"${quote_data['total_amount']:,.2f}"])
    
    # Create table
    product_table = Table(table_data, colWidths=[1.5*inch, 3*inch, 0.75*inch, 1.25*inch, 1.25*inch])
    
    # Apply table style
    table_style = TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(TURBO_BLUE)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows
        ('FONT', (0, 1), (-1, -2), 'Helvetica', 10),
        ('ALIGN', (2, 1), (-1, -2), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -2), 'RIGHT'),
        
        # Total row
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('ALIGN', (3, -1), (-1, -1), 'RIGHT'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F0F0F0')]),
    ])
    
    product_table.setStyle(table_style)
    elements.append(product_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Terms and Conditions
    elements.append(Paragraph("Terms and Conditions", heading_style))
    
    terms = [
        "• Prices are valid for 30 days from quote date",
        "• Delivery time: 2-4 weeks from order confirmation",
        "• Payment terms: Net 30 days",
        "• All prices are in USD and exclude taxes",
        "• Installation not included unless specified",
        "• Warranty: Standard manufacturer warranty applies",
        "• Shipping: FOB Origin unless otherwise specified"
    ]
    
    for term in terms:
        elements.append(Paragraph(term, styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_text = f"""
    <para align="center">
    <font color="{TURBO_BLUE}">Thank you for your business!</font><br/>
    For questions about this quote, please contact us at sales@turboair.com
    </para>
    """
    elements.append(Paragraph(footer_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    return output

def export_quote_to_excel(quote_data: Dict, items_df: pd.DataFrame, client_data: Dict, 
                         filename: str = None) -> str:
    """Export quote to Excel file and return filename"""
    if not filename:
        filename = f"Quote_{quote_data['quote_number']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    excel_buffer = generate_excel_quote(quote_data, items_df, client_data)
    
    # Save to file
    with open(filename, 'wb') as f:
        f.write(excel_buffer.read())
    
    return filename

def export_quote_to_pdf(quote_data: Dict, items_df: pd.DataFrame, client_data: Dict,
                       filename: str = None) -> str:
    """Export quote to PDF file and return filename"""
    if not filename:
        filename = f"Quote_{quote_data['quote_number']}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    pdf_buffer = generate_pdf_quote(quote_data, items_df, client_data)
    
    # Save to file
    with open(filename, 'wb') as f:
        f.write(pdf_buffer.read())
    
    return filename

def prepare_email_attachments(quote_data: Dict, items_df: pd.DataFrame, 
                            client_data: Dict) -> Dict[str, io.BytesIO]:
    """Prepare quote attachments for email"""
    attachments = {}
    
    # Generate Excel attachment
    excel_buffer = generate_excel_quote(quote_data, items_df, client_data)
    attachments[f"Quote_{quote_data['quote_number']}.xlsx"] = excel_buffer
    
    # Generate PDF attachment
    pdf_buffer = generate_pdf_quote(quote_data, items_df, client_data)
    attachments[f"Quote_{quote_data['quote_number']}.pdf"] = pdf_buffer
    
    return attachments