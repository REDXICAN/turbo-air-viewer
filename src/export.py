"""
Fixed Export functionality for Turbo Air Equipment Viewer
Improved PDF generation with better image handling and corruption fixes
"""

import io
import os
import base64
from datetime import datetime
from typing import Dict, List
import pandas as pd

# Excel exports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing import Image as XLImage
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF exports  
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    from PIL import Image as PILImage
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

def get_product_image_path(sku: str) -> str:
    """Get the path to product image with better path handling"""
    possible_paths = [
        f"pdf_screenshots/{sku}/{sku} P.1.png",
        f"pdf_screenshots/{sku}/{sku}_P.1.png", 
        f"pdf_screenshots/{sku}/{sku}.png",
        f"pdf_screenshots/{sku}/page_1.png",
        f"pdf_screenshots/{sku.upper()}/{sku.upper()} P.1.png",
        f"pdf_screenshots/{sku.upper()}/{sku.upper()}_P.1.png",
        f"pdf_screenshots/{sku.upper()}/{sku.upper()}.png",
        f"pdf_screenshots/{sku.lower()}/{sku.lower()} P.1.png",
        f"pdf_screenshots/{sku.lower()}/{sku.lower()}_P.1.png",
        f"pdf_screenshots/{sku.lower()}/{sku.lower()}.png"
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return path
    return None

def resize_image_for_pdf(image_path: str, max_width: float = 0.8, max_height: float = 0.6) -> io.BytesIO:
    """Resize image for PDF inclusion"""
    try:
        from PIL import Image as PILImage
        
        # Open and convert image
        with PILImage.open(image_path) as img:
            # Convert to RGB if necessary
            if img.mode in ('RGBA', 'LA', 'P'):
                img = img.convert('RGB')
            
            # Calculate new size maintaining aspect ratio
            img_width, img_height = img.size
            aspect_ratio = img_width / img_height
            
            # Convert inches to pixels (assuming 72 DPI)
            max_width_px = int(max_width * inch * 72 / 72)
            max_height_px = int(max_height * inch * 72 / 72)
            
            if img_width > max_width_px or img_height > max_height_px:
                if aspect_ratio > 1:  # Wider than tall
                    new_width = max_width_px
                    new_height = int(max_width_px / aspect_ratio)
                else:  # Taller than wide
                    new_height = max_height_px
                    new_width = int(max_height_px * aspect_ratio)
                
                img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # Save to BytesIO
            buffer = io.BytesIO()
            img.save(buffer, format='PNG', optimize=True, quality=85)
            buffer.seek(0)
            return buffer
    except Exception as e:
        print(f"Error resizing image {image_path}: {e}")
        return None

def generate_excel_quote(quote_data: Dict, items_df: pd.DataFrame, client_data: Dict) -> io.BytesIO:
    """Generate Excel quote with enhanced formatting and images - COMPLETELY FIXED"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl not available for Excel export")
    
    buffer = io.BytesIO()
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quote"
        
        # Styles
        header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        subheader_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color='20429C', end_color='20429C', fill_type='solid')
        light_fill = PatternFill(start_color='F2F2F7', end_color='F2F2F7', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = 'TURBO AIR - EQUIPMENT QUOTE'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Quote info
        current_row = 3
        ws[f'A{current_row}'] = 'Quote Information'
        ws[f'A{current_row}'].font = subheader_font
        current_row += 1
        
        quote_info = [
            ['Quote Number:', quote_data.get('quote_number', 'N/A')],
            ['Date:', datetime.now().strftime('%B %d, %Y')],
            ['Client:', client_data.get('company', 'N/A')],
            ['Contact:', client_data.get('contact_name', 'N/A')],
            ['Email:', client_data.get('contact_email', 'N/A')]
        ]
        
        for info in quote_info:
            ws[f'A{current_row}'] = info[0]
            ws[f'B{current_row}'] = info[1]
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
        
        current_row += 2
        
        # Items header
        ws[f'A{current_row}'] = 'Equipment List'
        ws[f'A{current_row}'].font = subheader_font
        current_row += 1
        
        # Table headers
        headers = ['Image', 'SKU', 'Description', 'Qty', 'Unit Price', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Image column
        ws.column_dimensions['B'].width = 20  # SKU
        ws.column_dimensions['C'].width = 40  # Description  
        ws.column_dimensions['D'].width = 8   # Qty
        ws.column_dimensions['E'].width = 12  # Unit Price
        ws.column_dimensions['F'].width = 12  # Total
        
        current_row += 1
        
        # Add items with images - FIXED VERSION
        for idx, item in items_df.iterrows():
            sku = str(item.get('sku', 'Unknown'))
            description = str(item.get('product_type', ''))
            quantity = int(item.get('quantity', 1))
            unit_price = float(item.get('price', 0))
            total_price = unit_price * quantity
            
            # Try to add product image with SAFE error handling
            image_added = False
            image_path = get_product_image_path(sku)
            if image_path and os.path.exists(image_path):
                try:
                    # Create image object with SAFE dimensions
                    img = XLImage(image_path)
                    
                    # Set FIXED safe dimensions
                    max_width = 80
                    max_height = 60
                    
                    # Calculate aspect ratio and resize
                    if hasattr(img, 'width') and hasattr(img, 'height'):
                        original_width = img.width
                        original_height = img.height
                        aspect_ratio = original_width / original_height
                        
                        if aspect_ratio > 1:  # Wider than tall
                            img.width = max_width
                            img.height = int(max_width / aspect_ratio)
                        else:  # Taller than wide
                            img.height = max_height
                            img.width = int(max_height * aspect_ratio)
                    else:
                        # Fallback dimensions
                        img.width = max_width
                        img.height = max_height
                    
                    # Add image to worksheet
                    ws.add_image(img, f'A{current_row}')
                    ws.row_dimensions[current_row].height = max(50, img.height + 10)
                    image_added = True
                    
                except Exception as img_error:
                    print(f"Could not add image for {sku}: {img_error}")
                    # Don't break - continue without image
            
            # If no image was added, set placeholder and row height
            if not image_added:
                ws[f'A{current_row}'] = '📷'
                ws.row_dimensions[current_row].height = 50
            
            # Add item data
            ws[f'B{current_row}'] = sku
            ws[f'C{current_row}'] = description
            ws[f'D{current_row}'] = quantity
            ws[f'E{current_row}'] = f"${unit_price:,.2f}"
            ws[f'F{current_row}'] = f"${total_price:,.2f}"
            
            # Format cells
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        # Totals
        current_row += 1
        subtotal = float(quote_data.get('subtotal', 0))
        tax_rate = float(quote_data.get('tax_rate', 0))
        # Handle tax_rate format - if it's already a percentage, don't multiply
        if tax_rate > 1:
            tax_rate_display = tax_rate
        else:
            tax_rate_display = tax_rate * 100
        tax_amount = float(quote_data.get('tax_amount', 0))
        total = float(quote_data.get('total_amount', 0))
        
        ws[f'E{current_row}'] = 'Subtotal:'
        ws[f'F{current_row}'] = f"${subtotal:,.2f}"
        ws[f'E{current_row}'].font = Font(bold=True)
        current_row += 1
        
        ws[f'E{current_row}'] = f'Tax ({tax_rate_display:.1f}%):'
        ws[f'F{current_row}'] = f"${tax_amount:,.2f}"
        ws[f'E{current_row}'].font = Font(bold=True)
        current_row += 1
        
        ws[f'E{current_row}'] = 'TOTAL:'
        ws[f'F{current_row}'] = f"${total:,.2f}"
        ws[f'E{current_row}'].font = Font(bold=True, size=14)
        ws[f'F{current_row}'].font = Font(bold=True, size=14)
        
        # Save to buffer with PROPER error handling
        try:
            wb.save(buffer)
            buffer.seek(0)
            
            # Verify the buffer has content
            if buffer.getvalue():
                return buffer
            else:
                raise Exception("Generated file is empty")
                
        except Exception as save_error:
            print(f"Error saving Excel file: {save_error}")
            raise
    
    except Exception as e:
        print(f"Error generating Excel: {e}")
        # Create a simple fallback Excel file
        buffer.seek(0)
        buffer.truncate(0)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = "Turbo Air Equipment Quote"
            ws['A2'] = f"Quote Number: {quote_data.get('quote_number', 'N/A')}"
            ws['A3'] = f"Error: {str(e)}"
            ws['A4'] = "Please contact support for assistance."
            wb.save(buffer)
            buffer.seek(0)
        except Exception as fallback_error:
            print(f"Fallback Excel creation failed: {fallback_error}")
        
        return buffer

def generate_pdf_quote(quote_data: Dict, items_df: pd.DataFrame, client_data: Dict) -> io.BytesIO:
    """Generate PDF quote with enhanced formatting and product images - COMPLETELY FIXED"""
    if not PDF_AVAILABLE:
        raise ImportError("reportlab not available for PDF export")
    
    buffer = io.BytesIO()
    
    try:
        # Create document with SAFE settings
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter, 
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#20429C'),
            spaceAfter=20,
            alignment=1  # Center
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'], 
            fontSize=14,
            textColor=colors.HexColor('#20429C'),
            spaceAfter=10
        )
        
        story = []
        
        # Title
        story.append(Paragraph("TURBO AIR - EQUIPMENT QUOTE", title_style))
        story.append(Spacer(1, 20))
        
        # Quote information table
        quote_info_data = [
            ['Quote Number:', str(quote_data.get('quote_number', 'N/A'))],
            ['Date:', datetime.now().strftime('%B %d, %Y')],
            ['Client:', str(client_data.get('company', 'N/A'))],
            ['Contact:', str(client_data.get('contact_name', 'N/A'))],
            ['Email:', str(client_data.get('contact_email', 'N/A'))]
        ]
        
        quote_table = Table(quote_info_data, colWidths=[2*inch, 3*inch])
        quote_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F2F2F7'))
        ]))
        
        story.append(quote_table)
        story.append(Spacer(1, 20))
        
        # Equipment list header
        story.append(Paragraph("Equipment List", header_style))
        
        # Prepare items data with SAFE image handling
        items_data = [['Image', 'SKU', 'Description', 'Qty', 'Unit Price', 'Total']]
        
        for idx, item in items_df.iterrows():
            sku = str(item.get('sku', 'Unknown'))
            description = str(item.get('product_type', ''))[:40]  # Truncate long descriptions
            quantity = int(item.get('quantity', 1))
            unit_price = float(item.get('price', 0))
            total_price = unit_price * quantity
            
            # Try to add product image with SAFE error handling
            image_cell = "📷"  # Default placeholder
            
            image_path = get_product_image_path(sku)
            if image_path and os.path.exists(image_path):
                try:
                    # Try with PIL resizing first
                    resized_buffer = resize_image_for_pdf(image_path, 0.6, 0.4)
                    if resized_buffer:
                        # Create ReportLab image from buffer
                        img = RLImage(resized_buffer, width=0.6*inch, height=0.4*inch)
                        image_cell = img
                    else:
                        # Fallback to direct file with smaller size
                        img = RLImage(image_path, width=0.5*inch, height=0.35*inch)
                        image_cell = img
                        
                except Exception as img_error:
                    print(f"Could not add image for {sku}: {img_error}")
                    # Keep default placeholder
            
            items_data.append([
                image_cell,
                sku,
                description,
                str(quantity),
                f"${unit_price:,.2f}",
                f"${total_price:,.2f}"
            ])
        
        # Create items table with SAFE styling
        try:
            items_table = Table(items_data, colWidths=[0.8*inch, 1.2*inch, 2.5*inch, 0.5*inch, 1*inch, 1*inch])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#20429C')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
                ('LEADING', (0, 0), (-1, -1), 12)
            ]))
            
            story.append(items_table)
        except Exception as table_error:
            print(f"Error creating items table: {table_error}")
            # Add simple text fallback
            story.append(Paragraph("Equipment items could not be displayed properly.", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Totals
        subtotal = float(quote_data.get('subtotal', 0))
        tax_rate = float(quote_data.get('tax_rate', 0))
        # Handle tax_rate format - if it's already a percentage, don't multiply
        if tax_rate > 1:
            tax_rate_display = tax_rate
        else:
            tax_rate_display = tax_rate * 100
        tax_amount = float(quote_data.get('tax_amount', 0))
        total = float(quote_data.get('total_amount', 0))
        
        totals_data = [
            ['Subtotal:', f"${subtotal:,.2f}"],
            [f'Tax ({tax_rate_display:.1f}%):', f"${tax_amount:,.2f}"],
            ['TOTAL:', f"${total:,.2f}"]
        ]
        
        totals_table = Table(totals_data, colWidths=[4*inch, 1.5*inch])
        totals_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 1), 10),
            ('FONTSIZE', (0, 2), (-1, 2), 12),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LINEBELOW', (0, 1), (-1, 1), 1, colors.black),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#20429C')),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.whitesmoke)
        ]))
        
        story.append(totals_table)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = """
        <para align="center" fontSize="9" textColor="#666666">
        Thank you for choosing Turbo Air Equipment!<br/>
        This quote is valid for 30 days from the date of issue.<br/>
        For questions, please contact us at turboairquotes@gmail.com
        </para>
        """
        story.append(Paragraph(footer_text, styles['Normal']))
        
        # Build PDF with SAFE error handling
        doc.build(story)
        
        # Verify the buffer has content
        buffer.seek(0)
        if not buffer.getvalue():
            raise Exception("Generated PDF is empty")
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"Error building PDF: {e}")
        # Create a simple text-only fallback PDF
        buffer.seek(0)
        buffer.truncate(0)
        
        try:
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = [
                Paragraph("Turbo Air Equipment Quote", getSampleStyleSheet()['Title']),
                Paragraph(f"Quote Number: {quote_data.get('quote_number', 'N/A')}", getSampleStyleSheet()['Normal']),
                Paragraph(f"Error generating full PDF: {str(e)}", getSampleStyleSheet()['Normal']),
                Paragraph("Please contact support for assistance.", getSampleStyleSheet()['Normal'])
            ]
            doc.build(story)
            buffer.seek(0)
        except Exception as fallback_error:
            print(f"Fallback PDF creation failed: {fallback_error}")
    
    return buffer

def prepare_email_attachments(quote_data: Dict, items_df: pd.DataFrame, client_data: Dict) -> Dict[str, io.BytesIO]:
    """Prepare email attachments (Excel and PDF) with better error handling"""
    attachments = {}
    
    try:
        # Generate Excel attachment
        if EXCEL_AVAILABLE:
            excel_buffer = generate_excel_quote(quote_data, items_df, client_data)
            if excel_buffer and excel_buffer.getvalue():  # Check if buffer has content
                attachments[f"Quote_{quote_data['quote_number']}.xlsx"] = excel_buffer
        else:
            print("Excel export not available - openpyxl not installed")
        
        # Generate PDF attachment
        if PDF_AVAILABLE:
            pdf_buffer = generate_pdf_quote(quote_data, items_df, client_data)
            if pdf_buffer and pdf_buffer.getvalue():  # Check if buffer has content
                attachments[f"Quote_{quote_data['quote_number']}.pdf"] = pdf_buffer
        else:
            print("PDF export not available - reportlab not installed")
        
    except Exception as e:
        print(f"Error preparing attachments: {e}")
        # Return any successfully created attachments
    
    return attachments